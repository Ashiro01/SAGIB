# en bienes_app/excel_generator.py

from io import BytesIO
import pandas as pd
from django.conf import settings
import os
from django.utils import timezone

# Importaciones necesarias de openpyxl para diseño
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# El nombre de la función ahora acepta un título para hacerlo dinámico
def generar_reporte_inventario_excel(bienes_queryset, titulo_reporte):
    # 1. Preparar los datos y crear el DataFrame de pandas
    datos_para_df = []
    for i, bien in enumerate(bienes_queryset):
        datos_para_df.append({
            'N°': i + 1,
            'CÓDIGO PATRIMONIAL': bien.codigo_patrimonial,
            'DESCRIPCIÓN': bien.descripcion,
            'CATEGORÍA': bien.categoria.nombre if bien.categoria else 'N/A',
            'MARCA': bien.marca or '',
            'MODELO': bien.modelo or '',
            'SERIAL': bien.serial or '',
            'UNIDAD ASIGNADA': bien.unidad_administrativa_actual.nombre if bien.unidad_administrativa_actual else 'N/A',
            'ESTADO': bien.get_estado_bien_display(),
            'FECHA DE ADQUISICIÓN': bien.fecha_adquisicion,
            'VALOR (Bs.)': bien.valor_unitario_bs
        })
    df = pd.DataFrame(datos_para_df)

    # 2. Escribir el DataFrame en un buffer en memoria
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inventario', index=False, startrow=12)

        # Auto-ajustar el ancho de las columnas
        worksheet = writer.sheets['Inventario']
        for i, col in enumerate(df.columns):
            column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[chr(65 + i)].width = column_len

    # 3. Usar openpyxl para cargar el workbook desde el buffer y añadir los elementos de diseño
    buffer.seek(0)
    workbook = load_workbook(buffer)
    sheet = workbook.active

    # --- Lógica del Encabezado (Membrete y Logo) ---
    logo_path = os.path.join(settings.BASE_DIR, 'static/images/logo_ipsfa.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.width = 75
        logo.height = 90
        sheet.add_image(logo, 'C2')

    # Definir el texto del membrete
    membrete_text = (
        'REPÚBLICA BOLIVARIANA DE VENEZUELA\n'
        'MINISTERIO DEL PODER POPULAR PARA LA DEFENSA\n'
        'VICEMINISTERIO DE SERVICIOS, PERSONAL Y LOGÍSTICA\n'
        'DIRECCIÓN GENERAL DE EMPRESAS Y SERVICIOS\n'
        'INSTITUTO DE PREVISIÓN SOCIAL DE LA FUERZA ARMADA NACIONAL BOLIVARIANA\n'
        'GERENCIA DE FINANZAS\n'
        'UNIDAD DE BIENES PÚBLICOS'
    )

    # Fusionar celdas y añadir el membrete
    sheet.merge_cells('B1:K7')
    cell = sheet['B1']
    cell.value = membrete_text
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(bold=True, size=10)

    # --- Título del Reporte ---
    sheet.merge_cells('A9:K10')
    cell = sheet['A9']
    cell.value = titulo_reporte
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=14)

    # --- Información adicional del reporte ---
    info_row = 12
    sheet.merge_cells(f'A{info_row}:K{info_row}')
    cell = sheet[f'A{info_row}']
    cell.value = f"Total de Bienes: {bienes_queryset.count()} | Fecha de Generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=11)

    # --- Color azul en encabezado de la tabla ---
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_row = 14  # Porque la tabla inicia en startrow=12 (fila 13 en Excel)
    for col in range(1, len(df.columns) + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')

    # --- Pie de página (Firma) ---
    last_row = sheet.max_row + 4

    sheet.merge_cells(f'A{last_row}:K{last_row}')
    cell = sheet[f'A{last_row}']
    cell.value = "____________________________________"
    cell.alignment = Alignment(horizontal='center')

    sheet.merge_cells(f'A{last_row+1}:K{last_row+1}')
    cell = sheet[f'A{last_row+1}']
    cell.value = "MY. ANDELSON JOSE PINTO HERRERA"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

    sheet.merge_cells(f'A{last_row+2}:K{last_row+2}')
    cell = sheet[f'A{last_row+2}']
    cell.value = "JEFE DEL DEPARTAMENTO DE BIENES PÚBLICOS"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

    # Centrar el contenido de todas las filas de datos
    data_start_row = 15  # Porque el DataFrame se escribe en startrow=12 (fila 13 en Excel, 1-indexed)
    data_end_row = sheet.max_row
    for row in sheet.iter_rows(min_row=data_start_row, max_row=data_end_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 4. Guardar el workbook modificado en un nuevo buffer final
    final_buffer = BytesIO()
    workbook.save(final_buffer)
    final_buffer.seek(0)

    return final_buffer

def generar_reporte_desincorporados_excel(movimientos_queryset, titulo_reporte, fecha_desde=None, fecha_hasta=None):
    datos_para_df = []
    for i, movimiento in enumerate(movimientos_queryset):
        datos_para_df.append({
            'N°': i + 1,
            'CÓDIGO PATRIMONIAL': movimiento.bien.codigo_patrimonial,
            'DESCRIPCIÓN': movimiento.bien.descripcion,
            'UBICACIÓN DEL BIEN': movimiento.bien.ubicacion_fisica_especifica or '',
            'MARCA': movimiento.bien.marca or '',
            'MODELO': movimiento.bien.modelo or '',
            'N° SERIAL': movimiento.bien.serial or '',
        })
    df = pd.DataFrame(datos_para_df)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Desincorporados', index=False, startrow=14)
        worksheet = writer.sheets['Desincorporados']
        col_widths = [7, 22, 35, 28, 18, 18, 18]
        for i, col in enumerate(df.columns):
            worksheet.column_dimensions[chr(65 + i)].width = col_widths[i] if i < len(col_widths) else 18
    buffer.seek(0)
    workbook = load_workbook(buffer)
    sheet = workbook.active
    # --- Encabezado y Logo ---
    logo_path = os.path.join(settings.BASE_DIR, 'static/images/logo_ipsfa.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.width = 75
        logo.height = 90
        sheet.add_image(logo, 'B2')
    membrete_text = (
        'REPÚBLICA BOLIVARIANA DE VENEZUELA\n'
        'MINISTERIO DEL PODER POPULAR PARA LA DEFENSA\n'
        'VICEMINISTERIO DE SERVICIOS, PERSONAL Y LOGÍSTICA\n'
        'DIRECCIÓN GENERAL DE EMPRESAS Y SERVICIOS\n'
        'INSTITUTO DE PREVISIÓN SOCIAL DE LA FUERZA ARMADA NACIONAL BOLIVARIANA\n'
        'GERENCIA DE FINANZAS\n'
        'UNIDAD DE BIENES PÚBLICOS'
    )
    sheet.merge_cells('B1:H7')
    cell = sheet['B1']
    cell.value = membrete_text
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(bold=True, size=10)
    # --- Título del Reporte ---
    sheet.merge_cells('A9:H10')
    cell = sheet['A9']
    cell.value = titulo_reporte
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=14)
    # --- Rango de Fechas ---
    rango_fechas = None
    if fecha_desde and fecha_hasta:
        rango_fechas = f"Desde: {fecha_desde}   Hasta: {fecha_hasta}"
    elif fecha_desde:
        rango_fechas = f"Desde: {fecha_desde}"
    elif fecha_hasta:
        rango_fechas = f"Hasta: {fecha_hasta}"
    if rango_fechas:
        sheet.merge_cells('A12:H12')
        cell = sheet['A12']
        cell.value = rango_fechas
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=11)
    # --- Color azul claro en encabezado de la tabla ---
    header_fill = PatternFill(start_color='B7D6F8', end_color='B7D6F8', fill_type='solid')
    header_row = 15  # Porque la tabla inicia en startrow=14 (fila 15 en Excel)
    for col in range(1, len(df.columns) + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
    # --- Pie de página (Firma) ---
    last_row = sheet.max_row + 4
    sheet.merge_cells(f'A{last_row}:H{last_row}')
    cell = sheet[f'A{last_row}']
    cell.value = "____________________________________"
    cell.alignment = Alignment(horizontal='center')
    sheet.merge_cells(f'A{last_row+1}:H{last_row+1}')
    cell = sheet[f'A{last_row+1}']
    cell.value = "MY. ANDELSON JOSE PINTO HERRERA"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    sheet.merge_cells(f'A{last_row+2}:H{last_row+2}')
    cell = sheet[f'A{last_row+2}']
    cell.value = "JEFE DEL DEPARTAMENTO DE BIENES PÚBLICOS"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    # Centrar el contenido de todas las filas de datos
    data_start_row = 16
    data_end_row = sheet.max_row
    for row in sheet.iter_rows(min_row=data_start_row, max_row=data_end_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    final_buffer = BytesIO()
    workbook.save(final_buffer)
    final_buffer.seek(0)
    return final_buffer

def generar_reporte_traslados_excel(movimientos_queryset, titulo_reporte, fecha_desde=None, fecha_hasta=None):
    datos_para_df = []
    for i, movimiento in enumerate(movimientos_queryset):
        datos_para_df.append({
            'N°': i + 1,
            'FECHA DE TRASLADO': movimiento.fecha_movimiento.strftime('%d/%m/%Y'),
            'CODIGO PATRIMONIAL': movimiento.bien.codigo_patrimonial,
            'DESCRIPCION': movimiento.bien.descripcion,
            'UNIDAD ORIGEN': movimiento.unidad_origen.nombre if movimiento.unidad_origen else 'N/A',
            'UNIDAD DESTINO': movimiento.unidad_destino.nombre if movimiento.unidad_destino else 'N/A',
            'NUEVO RESPONSABLE': movimiento.responsable_nuevo_nombre or '',
        })

    df = pd.DataFrame(datos_para_df)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Traslados', index=False, startrow=14)
        worksheet = writer.sheets['Traslados']
        col_widths = [7, 22, 22, 35, 28, 28, 28]
        for i, col in enumerate(df.columns):
            worksheet.column_dimensions[chr(65 + i)].width = col_widths[i] if i < len(col_widths) else 18

    buffer.seek(0)
    workbook = load_workbook(buffer)
    sheet = workbook.active
    # --- Encabezado y Logo ---
    logo_path = os.path.join(settings.BASE_DIR, 'static/images/logo_ipsfa.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.width = 75
        logo.height = 90
        sheet.add_image(logo, 'B2')

    membrete_text = (
        'REPÚBLICA BOLIVARIANA DE VENEZUELA\n'
        'MINISTERIO DEL PODER POPULAR PARA LA DEFENSA\n'
        'VICEMINISTERIO DE SERVICIOS, PERSONAL Y LOGÍSTICA\n'
        'DIRECCIÓN GENERAL DE EMPRESAS Y SERVICIOS\n'
        'INSTITUTO DE PREVISIÓN SOCIAL DE LA FUERZA ARMADA NACIONAL BOLIVARIANA\n'
        'GERENCIA DE FINANZAS\n'
        'UNIDAD DE BIENES PÚBLICOS'
    )
    sheet.merge_cells('B1:H7')
    cell = sheet['B1']
    cell.value = membrete_text
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(bold=True, size=10)

    # --- Título del Reporte ---
    sheet.merge_cells('A9:H10')
    cell = sheet['A9']
    cell.value = titulo_reporte
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=14)

    # --- Rango de Fechas ---
    rango_fechas = None
    if fecha_desde and fecha_hasta:
        rango_fechas = f"Desde: {fecha_desde}   Hasta: {fecha_hasta}"
    elif fecha_desde:
        rango_fechas = f"Desde: {fecha_desde}"
    elif fecha_hasta:
        rango_fechas = f"Hasta: {fecha_hasta}"

    if rango_fechas:
        sheet.merge_cells('A12:H12')
        cell = sheet['A12']
        cell.value = rango_fechas
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=11)

    # --- Color azul claro en encabezado de la tabla ---
    header_fill = PatternFill(start_color='B7D6F8', end_color='B7D6F8', fill_type='solid')
    header_row = 15  # Porque la tabla inicia en startrow=14 (fila 15 en Excel)
    for col in range(1, len(df.columns) + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
    
    # --- Pie de página (Firma) ---
    last_row = sheet.max_row + 4
    sheet.merge_cells(f'A{last_row}:H{last_row}')
    cell = sheet[f'A{last_row}']
    cell.value = "____________________________________"
    cell.alignment = Alignment(horizontal='center')

    sheet.merge_cells(f'A{last_row+1}:H{last_row+1}')
    cell = sheet[f'A{last_row+1}']
    cell.value = "MY. ANDELSON JOSE PINTO HERRERA"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

    sheet.merge_cells(f'A{last_row+2}:H{last_row+2}')
    cell = sheet[f'A{last_row+2}']
    cell.value = "JEFE DEL DEPARTAMENTO DE BIENES PÚBLICOS"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    # Centrar el contenido de todas las filas de datos
    data_start_row = 16
    data_end_row = sheet.max_row
    for row in sheet.iter_rows(min_row=data_start_row, max_row=data_end_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    final_buffer = BytesIO()
    workbook.save(final_buffer)
    final_buffer.seek(0)
    return final_buffer

# --- NUEVA FUNCIÓN PARA REPORTE DE DEPRECIACIÓN ---
def generar_reporte_depreciacion_excel(bienes_con_depreciacion, titulo_reporte):
    datos_para_df = []
    for i, bien in enumerate(bienes_con_depreciacion):
        depreciacion_acumulada = bien.ultima_depreciacion_acumulada or 0
        valor_neto = bien.ultimo_valor_neto or bien.valor_unitario_bs
        porcentaje_depreciacion = (depreciacion_acumulada / bien.valor_unitario_bs * 100) if bien.valor_unitario_bs > 0 else 0
        categoria_nombre = bien.categoria.nombre if bien.categoria else 'N/A'
        
        datos_para_df.append({
            'N°': i + 1,
            'CÓDIGO PATRIMONIAL': bien.codigo_patrimonial,
            'DESCRIPCIÓN': bien.descripcion,
            'CATEGORÍA': categoria_nombre,
            'VALOR ORIGINAL (Bs.)': bien.valor_unitario_bs,
            'DEPRECIACIÓN ACUMULADA (Bs.)': depreciacion_acumulada,
            'VALOR NETO EN LIBROS (Bs.)': valor_neto,
            '% DEPRECIACIÓN': f"{porcentaje_depreciacion:.1f}%"
        })
    
    df = pd.DataFrame(datos_para_df)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Depreciacion_Acumulada', index=False, startrow=12)
        worksheet = writer.sheets['Depreciacion_Acumulada']
        
        # Auto-ajustar el ancho de las columnas
        for i, col in enumerate(df.columns):
            column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[chr(65 + i)].width = column_len

    buffer.seek(0)
    workbook = load_workbook(buffer)
    sheet = workbook.active

    # --- Encabezado y Logo ---
    logo_path = os.path.join(settings.BASE_DIR, 'static/images/logo_ipsfa.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.width = 75
        logo.height = 90
        sheet.add_image(logo, 'C2')

    membrete_text = (
        'REPÚBLICA BOLIVARIANA DE VENEZUELA\n'
        'MINISTERIO DEL PODER POPULAR PARA LA DEFENSA\n'
        'VICEMINISTERIO DE SERVICIOS, PERSONAL Y LOGÍSTICA\n'
        'DIRECCIÓN GENERAL DE EMPRESAS Y SERVICIOS\n'
        'INSTITUTO DE PREVISIÓN SOCIAL DE LA FUERZA ARMADA NACIONAL BOLIVARIANA\n'
        'GERENCIA DE FINANZAS\n'
        'UNIDAD DE BIENES PÚBLICOS'
    )
    sheet.merge_cells('B1:H7')
    cell = sheet['B1']
    cell.value = membrete_text
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(bold=True, size=10)

    # --- Título del Reporte ---
    sheet.merge_cells('A9:H10')
    cell = sheet['A9']
    cell.value = titulo_reporte
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=14)

    # --- Información adicional del reporte ---
    info_row = 12
    sheet.merge_cells(f'A{info_row}:H{info_row}')
    cell = sheet[f'A{info_row}']
    cell.value = f"Total de Bienes: {bienes_con_depreciacion.count()} | Fecha de Generación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, size=11)

    # --- Color azul en encabezado de la tabla ---
    header_fill = PatternFill(start_color='4169E1', end_color='4169E1', fill_type='solid')
    header_row = 14
    for col in range(1, len(df.columns) + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')

    # --- Pie de página (Firma) ---
    last_row = sheet.max_row + 4
    sheet.merge_cells(f'A{last_row}:H{last_row}')
    cell = sheet[f'A{last_row}']
    cell.value = "____________________________________"
    cell.alignment = Alignment(horizontal='center')

    sheet.merge_cells(f'A{last_row+1}:H{last_row+1}')
    cell = sheet[f'A{last_row+1}']
    cell.value = "MY. ANDELSON JOSE PINTO HERRERA"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

    sheet.merge_cells(f'A{last_row+2}:H{last_row+2}')
    cell = sheet[f'A{last_row+2}']
    cell.value = "JEFE DEL DEPARTAMENTO DE BIENES PÚBLICOS"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

    # Centrar el contenido de todas las filas de datos
    data_start_row = 15
    data_end_row = sheet.max_row
    for row in sheet.iter_rows(min_row=data_start_row, max_row=data_end_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    final_buffer = BytesIO()
    workbook.save(final_buffer)
    final_buffer.seek(0)
    return final_buffer